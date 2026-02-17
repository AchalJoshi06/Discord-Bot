"""Donation tracking: lifetime stats from achievements and monthly snapshots."""
from typing import Dict, Any, Optional, List
from datetime import datetime, timezone
import calendar

from storage import (
    load_donation_snapshots, save_donation_snapshots, get_latest_snapshot
)

# Optional Excel export support (openpyxl)
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False


# Achievement names that contain lifetime donation stats
ACHIEVEMENT_NAMES = {
    "Friend in Need": "troops_donated",
    "Sharing is Caring": "spells_donated",
    "Siege Sharer": "siege_donated",
}


def extract_lifetime_donations(player_json: Dict[str, Any]) -> Dict[str, int]:
    """
    Extract lifetime donation statistics from player achievements.
    
    Returns dict with:
        - troops_donated: Lifetime troops donated
        - spells_donated: Lifetime spells donated
        - siege_donated: Lifetime siege machines donated
        - total_donated: Sum of all lifetime donations
    """
    achievements = player_json.get("achievements", [])
    if not isinstance(achievements, list):
        return {
            "troops_donated": 0,
            "spells_donated": 0,
            "siege_donated": 0,
            "total_donated": 0
        }
    
    lifetime = {
        "troops_donated": 0,
        "spells_donated": 0,
        "siege_donated": 0,
        "total_donated": 0
    }
    
    for achievement in achievements:
        name = achievement.get("name", "")
        value = achievement.get("value", 0)
        
        # Map achievement names to donation types
        # Try exact match first
        if name in ACHIEVEMENT_NAMES:
            key = ACHIEVEMENT_NAMES[name]
            try:
                lifetime[key] = int(value)
            except (ValueError, TypeError):
                pass
        # Also try case-insensitive match (some APIs return different casing)
        else:
            name_lower = name.lower()
            for ach_name, key in ACHIEVEMENT_NAMES.items():
                if ach_name.lower() == name_lower:
                    try:
                        lifetime[key] = int(value)
                    except (ValueError, TypeError):
                        pass
                    break
    
    lifetime["total_donated"] = (
        lifetime["troops_donated"] +
        lifetime["spells_donated"] +
        lifetime["siege_donated"]
    )
    
    return lifetime


def get_current_month_key() -> str:
    """Get current month in YYYY-MM format."""
    now = datetime.now(timezone.utc)
    return now.strftime("%Y-%m")


def create_donation_snapshot(
    clan_tag: str,
    members: List[Dict[str, Any]],
    player_data_cache: Dict[str, Dict[str, Any]]
) -> Dict[str, Any]:
    """
    Create a donation snapshot for the current month.
    
    Args:
        clan_tag: Clan tag
        members: List of clan members from API
        player_data_cache: Cache of player data (tag -> player_json)
    
    Returns:
        Snapshot dict with date and member donation data
    """
    month_key = get_current_month_key()
    snapshot = {
        "date": month_key,
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "members": {}
    }
    
    for member in members:
        tag = member.get("tag")
        if not tag:
            continue
        
        # Get player data (from cache or will be fetched)
        player = player_data_cache.get(tag)
        if not player:
            continue
        
        # Extract lifetime donations from achievements
        lifetime = extract_lifetime_donations(player)
        
        # Get seasonal donations (current season)
        seasonal = player.get("donations", 0)
        
        snapshot["members"][tag] = {
            "name": player.get("name", member.get("name", "Unknown")),
            "seasonal": int(seasonal),
            "lifetime": lifetime
        }
    
    return snapshot


def save_monthly_snapshot(
    clan_tag: str,
    snapshot: Dict[str, Any]
) -> bool:
    """Save a monthly snapshot to storage."""
    snapshots = load_donation_snapshots()
    
    if clan_tag not in snapshots:
        snapshots[clan_tag] = []
    
    # Check if snapshot for this month already exists
    month_key = snapshot["date"]
    clan_snapshots = snapshots[clan_tag]
    
    # Remove existing snapshot for this month if it exists
    snapshots[clan_tag] = [
        s for s in clan_snapshots if s.get("date") != month_key
    ]
    
    # Add new snapshot
    snapshots[clan_tag].append(snapshot)
    
    # Keep only last 24 months of snapshots
    sorted_snapshots = sorted(snapshots[clan_tag], key=lambda x: x.get("date", ""), reverse=True)
    snapshots[clan_tag] = sorted_snapshots[:24]
    
    success = save_donation_snapshots(snapshots)

    # If saving succeeded, also export to Excel (if available)
    if success:
        try:
            _write_snapshot_to_excel(clan_tag, snapshot)
        except Exception:
            # Don't fail saving if Excel export fails
            pass

    return success


def _write_snapshot_to_excel(clan_tag: str, snapshot: Dict[str, Any]) -> None:
    """Create or update an Excel workbook for donation history for a clan.

    Workbook layout:
      - Sheet 'Summary': Month | TotalMonthly | MemberCount
      - Sheet per month (YYYY-MM): Tag | Name | Monthly | Seasonal | Troops | Spells | Siege | LifetimeTotal
    """
    if not OPENPYXL_AVAILABLE:
        return

    month_key = snapshot.get("date")
    if not month_key:
        return

    # Calculate monthly donations (may return first-snapshot note)
    monthly_data = calculate_monthly_donations(clan_tag, month_key)

    # Build filename
    safe_tag = clan_tag.replace('#', '')
    fname = f"donation_history_{safe_tag}.xlsx"

    try:
        wb = load_workbook(fname)
    except Exception:
        wb = Workbook()

    # Summary sheet
    if "Summary" not in wb.sheetnames:
        ws_sum = wb.create_sheet("Summary")
        ws_sum.append(["Month", "TotalMonthly", "MemberCount"])
    else:
        ws_sum = wb["Summary"]

    # Update or append summary row for month
    found = False
    for row in ws_sum.iter_rows(min_row=2, values_only=False):
        cell_month = row[0]
        if str(cell_month.value) == month_key:
            # update
            row[1].value = monthly_data.get("total_monthly", 0) if monthly_data else 0
            row[2].value = len((monthly_data.get("members") if monthly_data else {}))
            found = True
            break
    if not found:
        ws_sum.append([month_key, monthly_data.get("total_monthly", 0) if monthly_data else 0, len((monthly_data.get("members") if monthly_data else {}))])

    # Create/replace month sheet
    if month_key in wb.sheetnames:
        std = wb[month_key]
        wb.remove(std)
    ws = wb.create_sheet(month_key)

    headers = ["Tag", "Name", "Monthly", "Seasonal", "Troops", "Spells", "Siege", "LifetimeTotal"]
    ws.append(headers)

    if monthly_data and monthly_data.get("members"):
        # Sort by monthly desc
        items = sorted(monthly_data["members"].items(), key=lambda x: x[1].get("monthly", 0), reverse=True)
        for tag, data in items:
            lifetime = data.get("lifetime", {})
            troops = lifetime.get("troops_donated", 0)
            spells = lifetime.get("spells_donated", 0)
            siege = lifetime.get("siege_donated", 0)
            lifetime_total = lifetime.get("total_donated", 0)
            ws.append([
                tag,
                data.get("name", "Unknown"),
                data.get("monthly", 0),
                data.get("seasonal", 0),
                troops,
                spells,
                siege,
                lifetime_total
            ])
    else:
        # First snapshot / no previous data: write seasonal counts as a snapshot
        members = snapshot.get("members", {})
        for tag, m in members.items():
            lifetime = m.get("lifetime", {})
            troops = lifetime.get("troops_donated", 0)
            spells = lifetime.get("spells_donated", 0)
            siege = lifetime.get("siege_donated", 0)
            total = lifetime.get("total_donated", 0)
            ws.append([
                tag,
                m.get("name", "Unknown"),
                m.get("seasonal", 0),
                m.get("seasonal", 0),
                troops,
                spells,
                siege,
                total
            ])

    # Auto-size columns (simple heuristic)
    for col_idx, col in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col:
            try:
                val = str(cell.value) if cell.value is not None else ""
                if len(val) > max_len:
                    max_len = len(val)
            except Exception:
                pass
        adjusted_width = (max_len + 2)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # Save workbook
    wb.save(fname)

def calculate_monthly_donations(
    clan_tag: str,
    month_key: Optional[str] = None
) -> Optional[Dict[str, Any]]:
    """
    Calculate donations made during a specific month.
    
    Args:
        clan_tag: Clan tag
        month_key: Month in YYYY-MM format (defaults to current month)
    
    Returns:
        Dict with month, member donations, and totals
    """
    if month_key is None:
        month_key = get_current_month_key()
    
    snapshots = load_donation_snapshots()
    clan_snapshots = snapshots.get(clan_tag, [])
    
    if not clan_snapshots:
        return None
    
    # Find snapshot for the target month
    target_snapshot = next(
        (s for s in clan_snapshots if s.get("date") == month_key),
        None
    )
    
    if not target_snapshot:
        return None
    
    # Find previous snapshot (to calculate difference)
    sorted_snapshots = sorted(clan_snapshots, key=lambda x: x.get("date", ""))
    target_index = next(
        (i for i, s in enumerate(sorted_snapshots) if s.get("date") == month_key),
        None
    )
    
    if target_index is None or target_index == 0:
        # First snapshot - can't calculate difference
        return {
            "month": month_key,
            "members": {},
            "total_monthly": 0,
            "note": "First snapshot - no previous data to compare"
        }
    
    prev_snapshot = sorted_snapshots[target_index - 1]
    
    # Calculate monthly donations (difference in seasonal counts)
    monthly_donations = {}
    total_monthly = 0
    
    target_members = target_snapshot.get("members", {})
    prev_members = prev_snapshot.get("members", {})
    
    for tag, data in target_members.items():
        target_seasonal = data.get("seasonal", 0)
        prev_data = prev_members.get(tag)
        
        if prev_data:
            prev_seasonal = prev_data.get("seasonal", 0)
            monthly = max(0, target_seasonal - prev_seasonal)
        else:
            # New member - use current seasonal as monthly
            monthly = target_seasonal
        
        monthly_donations[tag] = {
            "name": data.get("name", "Unknown"),
            "monthly": monthly,
            "seasonal": target_seasonal,
            "lifetime": data.get("lifetime", {})
        }
        total_monthly += monthly
    
    return {
        "month": month_key,
        "members": monthly_donations,
        "total_monthly": total_monthly
    }


def get_donation_history(clan_tag: str, limit: int = 12) -> List[Dict[str, Any]]:
    """
    Get donation history for a clan (last N months).
    
    Args:
        clan_tag: Clan tag
        limit: Number of months to return
    
    Returns:
        List of monthly donation summaries
    """
    snapshots = load_donation_snapshots()
    clan_snapshots = snapshots.get(clan_tag, [])
    
    if not clan_snapshots:
        return []
    
    # Sort by date (most recent first)
    sorted_snapshots = sorted(clan_snapshots, key=lambda x: x.get("date", ""), reverse=True)
    
    history = []
    for i, snapshot in enumerate(sorted_snapshots[:limit]):
        month_key = snapshot.get("date", "")
        
        # Calculate monthly donations
        monthly_data = calculate_monthly_donations(clan_tag, month_key)
        if monthly_data:
            history.append(monthly_data)
    
    return history


def get_player_donation_stats(
    tag: str,
    clan_tag: Optional[str] = None
) -> Optional[Dict[str, Any]]:
    """
    Get comprehensive donation stats for a player.
    
    Returns:
        Dict with lifetime, seasonal, monthly, and tracked totals
    """
    # Get latest snapshot
    if clan_tag:
        latest = get_latest_snapshot(clan_tag)
        if latest:
            member_data = latest.get("members", {}).get(tag)
            if member_data:
                return {
                    "lifetime": member_data.get("lifetime", {}),
                    "seasonal": member_data.get("seasonal", 0),
                    "snapshot_date": latest.get("date"),
                    "tracked_from": get_tracking_start_date(clan_tag)
                }
    
    return None


def get_tracking_start_date(clan_tag: str) -> Optional[str]:
    """Get the date when tracking started for a clan."""
    snapshots = load_donation_snapshots()
    clan_snapshots = snapshots.get(clan_tag, [])
    
    if not clan_snapshots:
        return None
    
    # Get earliest snapshot
    sorted_snapshots = sorted(clan_snapshots, key=lambda x: x.get("date", ""))
    if sorted_snapshots:
        return sorted_snapshots[0].get("date")
    
    return None

